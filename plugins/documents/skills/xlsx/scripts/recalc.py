"""Excel Formula Recalculation Script"""
import json
import os
import platform
import subprocess
import sys
from pathlib import Path
from office.soffice import get_soffice_env
from openpyxl import load_workbook

MACRO_DIR_MACOS = "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
MACRO_DIR_LINUX = "~/.config/libreoffice/4/user/basic/Standard"

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def setup_libreoffice_macro():
    macro_dir = os.path.expanduser(MACRO_DIR_MACOS if platform.system() == "Darwin" else MACRO_DIR_LINUX)
    macro_file = os.path.join(macro_dir, "Module1.xba")
    if os.path.exists(macro_file) and "RecalculateAndSave" in Path(macro_file).read_text():
        return True
    if not os.path.exists(macro_dir):
        subprocess.run(["soffice", "--headless", "--terminate_after_init"], capture_output=True, timeout=10, env=get_soffice_env())
        os.makedirs(macro_dir, exist_ok=True)
    try:
        Path(macro_file).write_text(RECALCULATE_MACRO)
        return True
    except Exception:
        return False


def recalc(filename, timeout=30):
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}
    abs_path = str(Path(filename).absolute())
    if not setup_libreoffice_macro():
        return {"error": "Failed to setup LibreOffice macro"}
    cmd = ["soffice", "--headless", "--norestore",
           "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application", abs_path]
    if platform.system() == "Linux":
        cmd = ["timeout", str(timeout)] + cmd
    result = subprocess.run(cmd, capture_output=True, text=True, env=get_soffice_env())
    if result.returncode != 0 and result.returncode != 124:
        return {"error": result.stderr or "Unknown error"}
    try:
        wb = load_workbook(filename, data_only=True)
        excel_errors = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
        error_details = {err: [] for err in excel_errors}
        total_errors = 0
        for sheet_name in wb.sheetnames:
            for row in wb[sheet_name].iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                error_details[err].append(f"{sheet_name}!{cell.coordinate}")
                                total_errors += 1
                                break
        wb.close()
        result_dict = {"status": "success" if total_errors == 0 else "errors_found", "total_errors": total_errors, "error_summary": {}}
        for err_type, locations in error_details.items():
            if locations:
                result_dict["error_summary"][err_type] = {"count": len(locations), "locations": locations[:20]}
        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = sum(1 for sn in wb_formulas.sheetnames for row in wb_formulas[sn].iter_rows() for cell in row if cell.value and isinstance(cell.value, str) and cell.value.startswith("="))
        wb_formulas.close()
        result_dict["total_formulas"] = formula_count
        return result_dict
    except Exception as e:
        return {"error": str(e)}


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        sys.exit(1)
    print(json.dumps(recalc(sys.argv[1], int(sys.argv[2]) if len(sys.argv) > 2 else 30), indent=2))
